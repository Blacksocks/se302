#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
Generates BOM and Pick'nPlace files for PCBPool :
  - BOM condensed with order codes (and link to distributor) to verify thta the component is in stock
  - Pick'n'Place file

This program needs:
  - two files generated from schematics by iCDB-to-BOM utility:
    - ../BOM-condensed.csv: a condensed BOM (csv, tab separated, with cdb2bom-condensed.asc parameter)
    - ../BOM-split.csv: a split BOM file (csv, tab separated, with cdb2bom-split.asc parameter)
  - four files generated from PCB by Report Writer utility (physical extractor):
    - ../PCB/vbreport/work/VBPCBP.TFI (fiducials)
    - ../PCB/vbreport/work/VBPCBP.TCM (components)
    - ../PCB/vbreport/work/VBPCBP.TCV (Part Number)
    - ../PCB/vbreport/work/VBPCBP.TCP (Part Number Property)
  - three templates of each output file : Template_Pick_Place.csv, Template_Fiducials.csv, Template_BOM.xslx
"""

import os, sys
import requests
from openpyxl import *
from bs4 import BeautifulSoup
from pprint import pprint

FIDUCIALS_FILENAME = "../PCB/vbreport/work/VBPCBP.TFI"
COMPONENTS_FILENAME = "../PCB/vbreport/work/VBPCBP.TCM"
PARTS_FILENAME = "../PCB/vbreport/work/VBPCBP.TCV"
PARTS_PROPS_FILENAME = "../PCB/vbreport/work/VBPCBP.TCP"

BOM_SPLIT_FILENAME = "../BOM-split.csv"
BOM_SPLIT_PARAMS_FILENAME = "../BOM-split.asc"

BOM_CONDENSED_FILENAME = "../BOM-condensed.csv"
BOM_CONDENSED_PARAMS_FILENAME = "../BOM-condensed.asc"

FIDUCIALS_OUTPUT_FILENAME = "Fiducials.txt"
PNP_OUTPUT_FILENAME = "PicknPlace.txt"
BOM_OUTPUT_FILENAME = "BOM.xlsx"

parts = {}

def build_fiducials():
    """
    Read PCB/vbreport/output/work/VBPCBP.TFI file.
    Return a dictionary of all fiducials.

    Fiducials output file has to follow this format:
    HEADER
    RefDesignator	X	Y	Rot	Value	Package	Side
    #Fuciducial1	63	32	0	1mm	round 1mm	BOTTOM
    #Fuciducial2	63	32	0	1mm	round 1mm	TOP

    Input file has this format:
    X           	Y           	Layer 	Rot         	PadID 	PadstackID	CompID	OrderCounter
    30.800      	3.800       	1     	0.000       	5     	5049      	-1    	1
    30.800      	3.800       	2     	0.000       	5     	5049      	-1    	2
    """
    with open(FIDUCIALS_FILENAME, encoding="iso8859-1") as fin:
        # Read fiducials from VB Report and output them
        vals = []
        lines = fin.readlines()
        for line in lines[1:]:
            line = [v.strip(" \t\r\n") for v in line.split("\t")]
            vals.append(line)

        # Extract to layer number from values
        top = min([v[2] for v in vals])
        bottom = max([v[2] for v in vals])

        # Replace layer number by layer name
        for v in vals:
            if v[2] == top:
                v[2] = "TOP"
            else:
                v[2] = "BOTTOM"

        # Generate output file
        with open(FIDUCIALS_OUTPUT_FILENAME, "w") as fout:
            fout.write("Filename: Fiducials.txt\n\nLage der Leiterplatte:\nuntere linke Ecke: X=0 / Y=0\n")
            fout.write("RefDesignator	X	Y	Rot	Value	Package	Side\n")
            for val in vals:
                line = ["FID%s"%val[7], val[0], val[1], val[3], "1mm", "round 1mm", val[2]]
                fout.write("\t".join(line) + "\n")

def read_components():
    """
    Read VBPCBP.TCM and VBPCBP.TCV and build a database of every component with
    all necessary fields. Return the components list.
    """
    comps = []

    with open(COMPONENTS_FILENAME, encoding="iso8859-1") as fin:
        """
        First read and parse VPBCPB.TCM which has the following format:
         CompID	RefDes	MountType	X	Y	ROT	Side	Mirror	Fixed	PartNumberID	CellID	PinCount
         3	R5	SMD	16.230	41.010	90.000	BOTTOM	YES	NO	26	11	2
         4	C14	SMD	18.130	41.736	270.000	BOTTOM	YES	NO	5	8	2
         5	R3	SMD	14.325	41.021	90.000	BOTTOM	YES	NO	26	11	2
        """

        lines = fin.readlines()
        for line in lines[1:]:
            fields = [v.strip(" \t\r\n") for v in line.split("\t")]
            comp = {}
            comp["RefDes"] = fields[1]
            comp["X"] = fields[3]
            comp["Y"] = fields[4]
            comp["Rot"] = fields[5]
            comp["Layer"] = fields[6]
            comp["PNID"] = fields[9]
            comp["CellID"] = fields[10]
            comps.append(comp)

        """
        Then read and parse VPBCPB.TCV which has the following format:
          PartNumberID	PartNumber	PartName	PartLabel	RefDesPre	TopCell	BottomCell	PartSource	Description
          23	00-0033	RES_0603_E24	RES_20_0603	R	R0603	R0603		Resistor20ohms/5%/case0603
          24	00-0074	RES_0603_E24	RES_1.0k_0603	R	R0603	R0603		Resistor1.0kohms/5%/case0603
        """
        with open(PARTS_FILENAME, encoding="iso8859-1") as fin:
            lines = fin.readlines()
            for line in lines[1:]:
                fields = [v.strip(" \t\r\n") for v in line.split("\t")]
                cl = [c for c in comps if c["PNID"] == fields[0]]
                for c in cl:
                    c["PartNumber"] = fields[1]
                    c["PartLabel"] = fields[3]
                    c["Cell"] = fields[5]
                    c["Description"] = fields[8]
                    c["Ordernumber"] = ""
                    c["Distributor"] = ""
                    c["Stock"] = ""
                    c["Weblink"] = ""
                    c["Description_distrib"] = ""

                    # If this PartNumber has not yet been seen, then add it to parts
                    PN = c["PartNumber"]
                    if PN not in parts.keys():
                        parts[PN] = {}


        """
        Then read and parse VPBCPB.TCP which has the following format:
          PropID	PartNumberID	PropertyName	PropertyValue
          1 	23	Type	Resistor
          2 	23	Value 	20
          3 	23	Tolerance 	5%
          4 	23	Power Dissipation 	0.0625
          5 	23	Cout HT 	0.045
          122 	8 	Type	Capacitor
          123 	8 	Value 	0.000022
          124 	8 	Tolerance 	20%
          125 	8 	Code Fournisseur	Fa : 2494270
          126 	8 	Cout HT 	0.206        """
        with open(PARTS_PROPS_FILENAME, encoding="iso8859-1") as fin:
            lines = fin.readlines()
            for line in lines[1:]:
                fields = [v.strip(" \t\r\n") for v in line.split("\t")]
                cl = [c for c in comps if c["PNID"] == fields[1]]
                for c in cl:
                    if fields[2] == 'Code Fournisseur':
                        parts[c["PartNumber"]]["Code Fournisseur"] = fields[3]
                    elif fields[2] == 'Value':
                        c["Value"] = fields[3]

        """
        Update official information from distributor : description, stock and URL.
        """
        fetch_distributor_infos()

        """
        Then read BOM_split.csv which has the following format :
          Ref Des	Quantity	Value	Part Label	Part Number	Description	Distributor code	Option	Cell	Placed
          R13	1	20 	RES_20_0603	00-0033	Resistor 20 ohms / 5% / case 0603	YES
          R31	1	20 	RES_20_0603	00-0033	Resistor 20 ohms / 5% / case 0603	NO
          R32	1	20 	RES_20_0603	00-0033	Resistor 20 ohms / 5% / case 0603	YES
          R8	1	1K 	RES_1.0k_0603	00-0074	Resistor 1.0k ohms / 5% / case 0603	Fa : 9331697	NO
        """

        with open(BOM_SPLIT_FILENAME, encoding="iso8859-1") as fin:
            lines = fin.readlines()
            for line in lines[1:]:
                fields = [v.strip(" \t\r\n") for v in line.split("\t")]
                if len(fields) != 9:
                    print("Continue pour ligne", line, " fields = ", fields)
                    continue
                cl = [c for c in comps if c["RefDes"] == fields[0]]
                if cl is None or len(cl) == 0:
                    print("Error : no ref des for component:", fields)
                    return
                if len(cl) > 1:
                    print("Error : multiple RefDes for different components:", cl)
                    return
                c = cl[0]
                c["Value"] = fields[2].strip()
                if c["Value"] == "":
                    c["Value"] = fields[3].strip()
                c["Place"] = fields[8].lower()
                if c["Place"].lower() != "yes" and c["Place"].lower() != "no":
                    c["Place"] = "no"

        """
        Then update Distributor, Stock and URL from parts database.
        """
        for c in comps:
            part = parts[c["PartNumber"]]
            c["Distributor"] = part["Distributor"]
            c["Stock"] = part["Stock"]
            c["Weblink"] = part["Weblink"]
            c["Description_distrib"] = part["Description_distrib"]
            c["Ordernumber"] = part["Ordernumber"]

        return comps

def fetch_distributor_infos():
    """
    For each part, extract from its "Code Fournisseur" the extract distributor
    in full letter and distributor code.
    For example : "Fa : 1234567" --> ("Farnell", "1234567")
    For Farnell, check stock.
    """
    for part in parts.values():
        if "Code Fournisseur" not in part.keys():
            (part["Ordernumber"], part["Distributor"], part["Description_distrib"],
                 part["Stock"], part["Weblink"]) = "", "", "", "", ""
            continue
        code = part["Code Fournisseur"]
        print("Looking up ", code)
        vals = code.split(":")
        if len(vals) != 2:
            print("Malformed distributor code for component", comp)
            (part["Ordernumber"], part["Distributor"], part["Description_distrib"],
                 part["Stock"], part["Weblink"]) = code, "???", "???", "???", "???"
            return

        distributor, code = vals

        distributor.strip()
        code.strip()

        part["Ordernumber"] = code

        if distributor.lower()[0:2] == "fa":
            part["Distributor"] = "Farnell"
        elif distributor.lower()[0:2] == "dg":
            part["Distributor"] = "Digikey France"
        elif distributor.lower()[0:2] == "mo":
            part["Distributor"] = "Mouser France"
        elif distributor.lower()[0:2] == "rs":
            part["Distributor"] = "Radiospares"

        # For Farnell, get the official description and stock
        if part["Distributor"] == "Farnell":
            part["Description_distrib"], part["Stock"], part["Weblink"] = get_Farnell_infos(code)
        else:
            part["Description_distrib"], part["Stock"], part["Weblink"] = "", "", ""


def get_Farnell_infos(code):
    """Get URL, description and availability of a component at Farnell.
    The query URL is http://fr.farnell.com/webapp/wcs/stores/servlet/Search?catalogId=15001&exaMfpn=true&mfpn=%s&categoryId=&langId=-2&searchRef=SearchLookAhead0&storeId=10160"%(code, )
    but most of the parameters seem to be ignored.
    TODO XXX : see if they are really usefull.
    This URL is then redirected on the component page.
    """
    url = "http://fr.farnell.com/webapp/wcs/stores/servlet/Search?exaMfpn=true&mfpn=%s"%(code, )
    #url = "http://fr.farnell.com/webapp/wcs/stores/servlet/Search?catalogId=15001&exaMfpn=true&mfpn=%s&categoryId=&langId=-2&searchRef=SearchLookAhead0&storeId=10160"%(code, )
    resp = requests.get(url)
    if resp.reason != "OK":
        print("Farnell site returned an error : status code = %d, reason = %s." % (
            resp.status_code, resp.reason))
        return "Error", "Error", "Error"
    # Find usefull part in HTML page
    root = BeautifulSoup(resp.content, 'html.parser')
    product = root.find("div", id="product")
    # Description
    description = product.h1.text + product.h2.find("span", itemprop="name").text
    print("Description=", description)
    # Disponibility
    availability = root.find("div", class_="avalabilityContainer")
    dispo = availability.find("p", class_="available")
    if dispo != None:
        dispo = "".join(dispo.text.strip().strip("En stock").split())
    else:
        dispo = "Plus stocké"
    #else:
    #    dispo = availability.dd
    #    dispo = dispo.text.strip()
    print("Dispo=", dispo)
    return description, dispo, url

def build_PNP(comps):
    """
    Build the PNP file according to this template :

      Filename: *FILENAME*

      Lage der Leiterplatte:
      untere linke Ecke: X=0 / Y=0

      RefDesignator	X	Y	Rot	Value	Package	Side
      #R48	7.537	20.53	0	0	R0603	TOP
      #R45	5.116	19.994	270	10	R0603	BOTTOM
      #R10	32.721	38.729	90	200	R0603	TOP
      #C14	34.959	27.655	90	100nF	C0603	BOTTOM
      #C16	12.456	7.079	180	100nF	C0603	TOP
      #D5	5.598	50.8	90	MMSD4148T1G	SOD123	BOTTOM
    """
    # Generate output file
    with open(PNP_OUTPUT_FILENAME, "w") as fout:
        fout.write("Filename: PicknPlace.txt\n\nLage der Leiterplatte:\nuntere linke Ecke: X=0 / Y=0\n")
        fout.write("RefDesignator	X	Y	Rot	Value	Package	Side\n")
        for comp in comps:
            line = [comp["RefDes"], comp["X"], comp["Y"], comp["Rot"], comp["Value"], comp["Cell"], comp["Layer"]]
            fout.write("\t".join(line) + "\n")

def generate_BOM(comps):
    """
    Build the BOM for PCBPool.
    Format of the BOM : first line is the columns header with the following meaning:

    Part : RefDes
    Value : Value
    Device : Part Label
    Package : Cell
    Description : Description
    Description2 : Description from distributor (MUST BE MANUALLY REMOVED)
    Qty : Number per board
    Place : MUST BE MANUALLY FILLED IN (YES/NO)
    Provided_by_customer : MUST BE MANUALLY FILLED IN (YES/NO)
    Distributor : Farnell / Mouser / ...
    Ordernumber : Code distributeur (Ordernumber in this script)
    Weblink : URL at the distributor
    Remarks_customer : remarks
    Unit_price
    Total_price
    REMARKS_BETA
    OPTION1
    OPTION2
    OPTION3

    The BOM must be condensed.
    """

    # Create workbook / get first worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "ORDER"

    # Header
    headers = ["Part", "Value", "Device", "Package", "Description", "Description2",
               "Qty", "Place", "Provided_by_customer", "Distributor", "Ordernumber",
               "Weblink", "Remarks_customer", "Unit_price", "Total_price",
               "REMARKS_BETA", "OPTION1", "OPTION2", "OPTION3"]
    for i, h in enumerate(headers):
        c = ws.cell(row=1, column=i+1)
        c.value = h
        c.fill = styles.PatternFill("solid", fgColor='ccffcc')
        bd = styles.Side(style='thin', color="000000")
        c.border = styles.Border(top=bd, right=bd, bottom=bd, left=bd)

    # Lines
    fill = styles.PatternFill("solid", fgColor='ffffcc')
    bd = styles.Side(style='thin', color="000000")

    # Open condensed BOM
    row = 2
    with open(BOM_CONDENSED_FILENAME, encoding="iso8859-1") as fin:
        lines = fin.readlines()
        for line in lines[1:]:
            placed_comps =  []
            unplaced_comps =  []
            fields = [v.strip(" \t\r\n") for v in line.split("\t")]
            # Filter placed components from non placed ones
            placed_comps = [c for c in comps if c["RefDes"] in fields[0].split(",") and c["Place"].lower() == "yes"]
            unplaced_comps = [c for c in comps if c["RefDes"] in fields[0].split(",") and c["Place"].lower() != "yes"]
            # Fill in BOM file
            #"RefDes", "Value", "PartLabel", "Cell", "Description", "Description2",
            #"Qty", "Place", "Provided_by_customer", "Distributor", "Ordernumber",
            #"Weblink", "Remarks_customer", "Unit_price", "Total_price",
            #"REMARKS_BETA", "OPTION1", "OPTION2", "OPTION3"]
            if placed_comps != []:
                fill_BOM_line(ws, row, placed_comps, True)
                row = row + 1
            if unplaced_comps != []:
                fill_BOM_line(ws, row, unplaced_comps, False)
                row = row + 1

    # Adjust cell size
    dims = {}
    for row in ws.rows:
        for cell in row:
            if cell.value:
                dims[cell.column] = max((dims.get(cell.column, 0), len(cell.value)))
    for col, value in dims.items():
        ws.column_dimensions[col].width = value

    # Save BOM file
    wb.save(BOM_OUTPUT_FILENAME)

def fill_BOM_line(ws, line, comp_list, placed):
    # RefDes
    cell = ws.cell(row=line, column=1)
    cell.value = ",".join([c["RefDes"] for c in comp_list])

    c = comp_list[0]

    # Value
    cell = ws.cell(row=line, column=2)
    cell.value = c["Value"]

    # PartLabel
    cell = ws.cell(row=line, column=3)
    cell.value = c["PartLabel"]

    # Cell
    cell = ws.cell(row=line, column=4)
    cell.value = c["Cell"]

    # Description
    cell = ws.cell(row=line, column=5)
    cell.value = c["Description"]

    # Description2
    cell = ws.cell(row=line, column=6)
    cell.value = c["Description_distrib"]

    # Qty
    cell = ws.cell(row=line, column=7)
    cell.value = str(len(comp_list))

    # Place
    cell = ws.cell(row=line, column=8)
    if placed :
        cell.value = "YES"
    else:
        cell.value = "NO"

    # Distributor
    cell = ws.cell(row=line, column=10)
    cell.value = c["Distributor"]

    # Ordernumber
    cell = ws.cell(row=line, column=11)
    cell.value = c["Ordernumber"]

    # Weblink
    cell = ws.cell(row=line, column=12)
    if len(c["Weblink"]) != 0 :
        cell.value = "Click here"
        cell.hyperlink = c["Weblink"]

    # Remarks_customer (available or not at distributor)
    cell = ws.cell(row=line, column=13)
    cell.value = c["Stock"]

    # PN to ease library update
    cell = ws.cell(row=line, column=14)
    cell.value = c["PartNumber"]

comps = read_components()
#pprint(comps)
build_fiducials()
build_PNP(comps)
generate_BOM(comps)
